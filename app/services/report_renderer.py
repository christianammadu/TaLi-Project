"""Render financial statements as branded PDF (ReportLab) or spreadsheet (openpyxl).

D-01 "Market Ledger" rebrand (WP-10). Print adaptation: documents use WHITE paper
(legibility + ink saving); terracotta is kept only for the header rule, section
underlines, total rules and the wordmark dot. Pure rendering — takes already-queried
data + metadata, writes a file into a temp directory, returns file descriptors. The
caller deletes ``meta['tmpdir']`` afterwards.

Three statement types, one engine:
  * ``income_statement`` — Profit & Loss (revenue − COGS = gross; − opex = net).
  * ``transactions``     — Statement of Account ledger (Money in / out / running Balance).
  * ``cashflow``         — monthly inflow/outflow/net/closing per currency.

Pure-Python deps (reportlab, openpyxl) — install on the PythonAnywhere free tier with
no system libraries.
"""

import os
import re
import tempfile
from datetime import datetime

from reportlab.lib import colors
from reportlab.lib.colors import HexColor
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable,
)

PDF_MIME = "application/pdf"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# ---------------------------------------------------------------- brand tokens

class Brand:
    """Print-adapted D-01 palette (mirrors designs/.../statement-render-spec.md)."""
    INK = HexColor('#241F1A')        # body text
    SOFT = HexColor('#6B6258')       # labels / captions
    ACCENT = HexColor('#C2562F')     # header rule · section heads · total rules · dot
    LINE = HexColor('#D9D2C6')       # table rules
    LINE_SOFT = HexColor('#ECE6DB')  # row separators
    ZEBRA = HexColor('#FCFAF6')      # even-row zebra
    BAND = HexColor('#FBF7EF')       # table header fill
    POS = HexColor('#1E7A45')        # net-positive figures
    NEG = HexColor('#B23A2E')        # negatives / money-out

_NEG_HEX = '#B23A2E'
_POS_HEX = '#1E7A45'

# ---------------------------------------------------------------- fonts

_FONT_DIR = os.path.join(os.path.dirname(__file__), '..', 'static', 'fonts')

FONTS_OK = False
F_DISPLAY = 'Times-Bold'        # headers / wordmark (fallback)
F_BODY = 'Helvetica'            # body / labels (fallback)
F_BODY_SB = 'Helvetica-Bold'    # semibold body (fallback)
F_SYMBOL = 'Helvetica'          # font carrying the ₦ glyph (fallback has none → see _naira)


def _register_fonts():
    """Register the brand TTFs once; fall back to the standard PDF fonts if absent.

    Hanken Grotesk has no ₦ glyph, but Fraunces does — so the Naira symbol is drawn
    in Fraunces (F_SYMBOL) inline while figures stay in the body font.
    """
    global FONTS_OK, F_DISPLAY, F_BODY, F_BODY_SB, F_SYMBOL
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        pdfmetrics.registerFont(TTFont('Fraunces', os.path.join(_FONT_DIR, 'Fraunces-SemiBold.ttf')))
        pdfmetrics.registerFont(TTFont('Hanken', os.path.join(_FONT_DIR, 'HankenGrotesk-Regular.ttf')))
        pdfmetrics.registerFont(TTFont('Hanken-SB', os.path.join(_FONT_DIR, 'HankenGrotesk-SemiBold.ttf')))
        F_DISPLAY, F_BODY, F_BODY_SB, F_SYMBOL = 'Fraunces', 'Hanken', 'Hanken-SB', 'Fraunces'
        FONTS_OK = True
    except Exception as e:  # pragma: no cover - depends on bundled assets
        print(f"[report_renderer] brand fonts unavailable, using Helvetica/Times fallback: {e}")
        FONTS_OK = False


_register_fonts()

# ---------------------------------------------------------------- money

_CCY = {'NGN': '₦', 'USD': '$', 'GBP': '£', 'EUR': '€', 'GHS': '₵', 'KES': 'KSh ', 'ZAR': 'R'}


def _symbol(currency):
    # Plain-text symbol — always the real glyph (₦/$/£…); fine in chat text + Excel.
    cur = (currency or 'NGN').upper()
    return _CCY.get(cur, cur + ' ')


def format_money(amount, currency='NGN'):
    """Plain-text money: ``₦1,234.00`` / ``(₦98,000.00)`` for negatives. Used for
    chat summaries and captions. Excel writes raw numbers + a number format."""
    sym = _symbol(currency)
    neg = float(amount) < 0
    s = f"{sym}{abs(float(amount)):,.2f}"
    return f"({s})" if neg else s


def _money_markup(amount, currency='NGN', color=None):
    """ReportLab Paragraph markup for a money figure — the ₦ symbol is rendered in
    Fraunces (which carries the glyph) while the figures use the body font."""
    cur = (currency or 'NGN').upper()
    neg = float(amount) < 0
    digits = f"{abs(float(amount)):,.2f}"
    if cur == 'NGN':
        # Hanken has no ₦; draw it in Fraunces when bundled, else the 'NGN ' fallback.
        sym = f'<font name="{F_SYMBOL}">₦</font>' if FONTS_OK else 'NGN '
    else:
        sym = _CCY.get(cur, cur + ' ')
    body = f"{sym}{digits}"
    if neg:
        body = f"({body})"
    if color:
        body = f'<font color="{color}">{body}</font>'
    return body


# ---------------------------------------------------------------- helpers

def _slug(text):
    return re.sub(r'[^A-Za-z0-9]+', '_', text or '').strip('_') or 'statement'


def _tmpdir(meta):
    if not meta.get('tmpdir'):
        meta['tmpdir'] = tempfile.mkdtemp(prefix='tali_report_')
    return meta['tmpdir']


def _group_rows_by_currency(rows):
    by_cur = {}
    for r in rows or []:
        by_cur.setdefault(r.get('currency') or 'NGN', []).append(r)
    return by_cur


def render(kind, data, meta, fmt='pdf'):
    """Render the requested format(s); return a list of {'path','filename','mime'}.

    ``kind`` is 'income_statement' | 'transactions' | 'cashflow'. ``fmt`` is
    'pdf' | 'xlsx' | 'both'. For income_statement/transactions ``data`` is the raw
    transaction rows; for cashflow it's the per-currency monthly buckets.
    """
    formats = ['pdf', 'xlsx'] if fmt == 'both' else [fmt]
    files = []
    for f in formats:
        if f == 'xlsx':
            files.append(_render_xlsx(kind, data, meta))
        else:
            files.append(_render_pdf(kind, data, meta))
    return files


# ============================================================ income-statement math

def compute_income_statement(rows):
    """Derive a per-currency P&L from raw transaction rows.

    Stock is not valued in the ledger, so COGS = purchases (opening = closing = 0)
    and the statement is flagged ``stock_valued=False``.
    """
    by_cur = {}
    for r in rows or []:
        cur = r.get('currency') or 'NGN'
        d = by_cur.setdefault(cur, {'revenue': {}, 'purchases': 0.0, 'expenses': {}})
        amt = float(r.get('amount') or 0)
        typ = r.get('type')
        act = (r.get('action') or '').lower()
        cat = r.get('category') or 'Miscellaneous'
        if typ == 'income':
            d['revenue'][cat] = d['revenue'].get(cat, 0.0) + amt
        else:
            if act == 'purchase':
                d['purchases'] += amt
            else:
                d['expenses'][cat] = d['expenses'].get(cat, 0.0) + amt
    out = {}
    for cur, d in by_cur.items():
        total_rev = sum(d['revenue'].values())
        cogs = d['purchases']
        gross = total_rev - cogs
        total_exp = sum(d['expenses'].values())
        out[cur] = {
            'revenue': sorted(d['revenue'].items(), key=lambda x: -x[1]),
            'total_revenue': total_rev,
            'purchases': d['purchases'],
            'cogs': cogs,
            'stock_valued': False,
            'gross_profit': gross,
            'expenses': sorted(d['expenses'].items(), key=lambda x: -x[1]),
            'total_expenses': total_exp,
            'net_profit': gross - total_exp,
        }
    return out


# ============================================================ PDF

_USABLE_W = A4[0] - 2 * 18 * mm    # content width inside 18mm margins


def _styles():
    return {
        'biz_name': ParagraphStyle('biz_name', fontName=F_DISPLAY, fontSize=14, textColor=Brand.INK, leading=17),
        'soft': ParagraphStyle('soft', fontName=F_BODY, fontSize=10, textColor=Brand.SOFT, leading=14),
        'soft_r': ParagraphStyle('soft_r', fontName=F_BODY, fontSize=10, textColor=Brand.SOFT, leading=14, alignment=2),
        'sec': ParagraphStyle('sec', fontName=F_DISPLAY, fontSize=12.5, textColor=Brand.ACCENT, leading=15, spaceBefore=2),
        'item': ParagraphStyle('item', fontName=F_BODY, fontSize=10.5, textColor=Brand.SOFT, leading=14),
        'item_ink': ParagraphStyle('item_ink', fontName=F_BODY_SB, fontSize=10.5, textColor=Brand.INK, leading=14),
        'total_l': ParagraphStyle('total_l', fontName=F_BODY_SB, fontSize=11, textColor=Brand.INK, leading=15),
        'grand_l': ParagraphStyle('grand_l', fontName=F_DISPLAY, fontSize=15, textColor=Brand.INK, leading=19),
        'm_item': ParagraphStyle('m_item', fontName=F_BODY, fontSize=10.5, textColor=Brand.INK, leading=14, alignment=2),
        'm_total': ParagraphStyle('m_total', fontName=F_BODY_SB, fontSize=11, textColor=Brand.INK, leading=15, alignment=2),
        'm_grand': ParagraphStyle('m_grand', fontName=F_DISPLAY, fontSize=15, textColor=Brand.INK, leading=19, alignment=2),
        # table
        'th': ParagraphStyle('th', fontName=F_BODY_SB, fontSize=8.5, textColor=Brand.SOFT, leading=11),
        'th_r': ParagraphStyle('th_r', fontName=F_BODY_SB, fontSize=8.5, textColor=Brand.SOFT, leading=11, alignment=2),
        'td': ParagraphStyle('td', fontName=F_BODY, fontSize=8.8, textColor=Brand.INK, leading=12),
        'td_r': ParagraphStyle('td_r', fontName=F_BODY, fontSize=8.8, textColor=Brand.INK, leading=12, alignment=2),
        'td_soft': ParagraphStyle('td_soft', fontName=F_BODY, fontSize=8.8, textColor=Brand.SOFT, leading=12),
        'tf': ParagraphStyle('tf', fontName=F_BODY_SB, fontSize=8.8, textColor=Brand.INK, leading=12),
        'tf_r': ParagraphStyle('tf_r', fontName=F_BODY_SB, fontSize=8.8, textColor=Brand.INK, leading=12, alignment=2),
        'note': ParagraphStyle('note', fontName=F_BODY, fontSize=8.5, textColor=Brand.SOFT, leading=12),
    }


def _section_head(text, st):
    """Section heading: Fraunces accent uppercase + 1px bottom rule."""
    return [
        Spacer(1, 8),
        Paragraph(text.upper(), st['sec']),
        HRFlowable(width='100%', thickness=1, color=Brand.LINE, spaceBefore=4, spaceAfter=4),
    ]


def _pl_line(label, amount, currency, st, kind='item'):
    """One P&L line as a 2-col table: label (left) + amount (right), styled by kind."""
    amount_w = 140
    if kind == 'indent':
        lab = Paragraph('&nbsp;&nbsp;&nbsp;&nbsp;' + label, st['item'])
        amt = Paragraph(_money_markup(amount, currency), st['m_item'])
        lines = []
    elif kind == 'sub':
        lab = Paragraph(label, st['item_ink'])
        amt = Paragraph(_money_markup(amount, currency), st['m_total'])
        lines = [('LINEABOVE', (0, 0), (-1, 0), 0.75, Brand.LINE_SOFT)]
    elif kind == 'total':
        lab = Paragraph(label, st['total_l'])
        amt = Paragraph(_money_markup(amount, currency), st['m_total'])
        lines = [('LINEABOVE', (0, 0), (-1, 0), 1.5, Brand.INK)]
    elif kind == 'grand':
        color = _POS_HEX if amount >= 0 else _NEG_HEX
        lab = Paragraph(label, st['grand_l'])
        amt = Paragraph(_money_markup(amount, currency, color=color), st['m_grand'])
        lines = [
            ('LINEABOVE', (0, 0), (-1, 0), 2, Brand.ACCENT),
            ('LINEBELOW', (0, 0), (-1, 0), 1.4, Brand.ACCENT),
        ]
    else:  # item
        lab = Paragraph(label, st['item_ink'])
        amt = Paragraph(_money_markup(amount, currency), st['m_item'])
        lines = []
    t = Table([[lab, amt]], colWidths=[_USABLE_W - amount_w, amount_w])
    t.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 4 if kind in ('item', 'indent') else 7),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4 if kind in ('item', 'indent') else 7),
    ] + lines))
    return t


def _income_story(rows, meta, st):
    by_cur = compute_income_statement(rows)
    story = []
    multi = len(by_cur) > 1
    for cur, p in sorted(by_cur.items()):
        if multi:
            story += [Spacer(1, 6), Paragraph(f"Currency: {cur}", st['biz_name'])]
        # Revenue
        story += _section_head('Revenue', st)
        if p['revenue']:
            for label, amt in p['revenue']:
                story.append(_pl_line(label, amt, cur, st, 'indent'))
        else:
            story.append(_pl_line('Sales', 0.0, cur, st, 'indent'))
        story.append(_pl_line('Total revenue', p['total_revenue'], cur, st, 'sub'))
        # COGS
        story += _section_head('Cost of goods sold', st)
        story.append(_pl_line('Add: Purchases', p['purchases'], cur, st, 'indent'))
        if not p['stock_valued']:
            story.append(Paragraph('Stock not valued — COGS taken as purchases for the period.', st['note']))
        story.append(_pl_line('Cost of goods sold', p['cogs'], cur, st, 'sub'))
        story.append(_pl_line('Gross profit', p['gross_profit'], cur, st, 'total'))
        # Operating expenses
        story += _section_head('Operating expenses', st)
        if p['expenses']:
            for label, amt in p['expenses']:
                story.append(_pl_line(label, amt, cur, st, 'indent'))
        else:
            story.append(Paragraph('No operating expenses recorded for this period.', st['note']))
        story.append(_pl_line('Total operating expenses', p['total_expenses'], cur, st, 'sub'))
        # Net profit
        story.append(Spacer(1, 2))
        story.append(_pl_line('Net profit', p['net_profit'], cur, st, 'grand'))
    return story


def _account_story(rows, meta, st):
    """Statement of Account — Date · Description · Cat · Money in · Money out · Balance."""
    by_cur = _group_rows_by_currency(rows)
    story = []
    for cur, crows in sorted(by_cur.items()):
        opening = float(meta.get('opening_balance', {}).get(cur, 0.0)) if isinstance(meta.get('opening_balance'), dict) else 0.0
        header = [Paragraph(h, st['th'] if i < 3 else st['th_r'])
                  for i, h in enumerate(['Date', 'Description', 'Cat.', 'Money in', 'Money out', 'Balance'])]
        table = [header]
        balance = opening
        table.append([Paragraph(_short_date(crows[0]['date']) if crows else '', st['td']),
                      Paragraph('Opening balance', st['td_soft']), Paragraph('—', st['td_soft']),
                      Paragraph('', st['td_r']), Paragraph('', st['td_r']),
                      Paragraph(_money_markup(balance, cur), st['td_r'])])
        tot_in = tot_out = 0.0
        for r in crows:
            amt = float(r['amount'])
            is_in = r.get('type') == 'income'
            if is_in:
                tot_in += amt
                balance += amt
            else:
                tot_out += amt
                balance -= amt
            desc = (r.get('item') or (r.get('action') or '').title() or '—')
            if r.get('action'):
                desc = f"{desc} — {r['action']}"
            table.append([
                Paragraph(_short_date(r['date']), st['td']),
                Paragraph(desc[:46], st['td']),
                Paragraph((r.get('category') or '—')[:16], st['td_soft']),
                Paragraph(_money_markup(amt, cur, color=_POS_HEX) if is_in else '', st['td_r']),
                Paragraph(_money_markup(amt, cur, color=_NEG_HEX) if not is_in else '', st['td_r']),
                Paragraph(_money_markup(balance, cur), st['td_r']),
            ])
        table.append([
            Paragraph('Totals', st['tf']), Paragraph('', st['tf']), Paragraph('', st['tf']),
            Paragraph(_money_markup(tot_in, cur), st['tf_r']),
            Paragraph(_money_markup(tot_out, cur), st['tf_r']),
            Paragraph(_money_markup(balance, cur), st['tf_r']),
        ])
        col_w = [48, _USABLE_W - 48 - 64 - 74 - 74 - 78, 64, 74, 74, 78]
        t = Table(table, colWidths=col_w, repeatRows=1)
        n = len(table)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), Brand.BAND),
            ('LINEBELOW', (0, 0), (-1, 0), 1.5, Brand.INK),
            ('ROWBACKGROUNDS', (0, 1), (-1, n - 2), [colors.white, Brand.ZEBRA]),
            ('LINEBELOW', (0, 1), (-1, n - 3), 0.5, Brand.LINE_SOFT),
            ('LINEABOVE', (0, n - 1), (-1, n - 1), 1.5, Brand.INK),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))
        if len(by_cur) > 1:
            story.append(Paragraph(f"Currency: {cur}", st['biz_name']))
            story.append(Spacer(1, 4))
        story.append(t)
        story.append(Spacer(1, 6))
        story.append(Paragraph(
            'Money in = receipts (credit) · Money out = payments (debit) · Balance is the running cash position.',
            st['note']))
        story.append(Spacer(1, 10))
    if not by_cur:
        story.append(Paragraph('No transactions found for this period.', st['note']))
    return story


def _cashflow_story(data, meta, st):
    """Cashflow — opening → monthly in/out/net → closing, per currency (direct, monthly)."""
    story = []
    for cur, rows in sorted((data or {}).items()):
        opening = 0.0
        header = [Paragraph(h, st['th'] if i == 0 else st['th_r'])
                  for i, h in enumerate(['Month', 'Cash in', 'Cash out', 'Net', 'Balance'])]
        table = [header]
        tot_in = tot_out = 0.0
        closing = opening
        for r in rows:
            tot_in += r['inflow']
            tot_out += r['outflow']
            closing = opening + r['cumulative']
            table.append([
                Paragraph(_month_label(r['month']), st['td']),
                Paragraph(_money_markup(r['inflow'], cur, color=_POS_HEX), st['td_r']),
                Paragraph(_money_markup(r['outflow'], cur, color=_NEG_HEX), st['td_r']),
                Paragraph(_money_markup(r['net'], cur), st['td_r']),
                Paragraph(_money_markup(closing, cur), st['td_r']),
            ])
        table.append([
            Paragraph('Totals', st['tf']),
            Paragraph(_money_markup(tot_in, cur), st['tf_r']),
            Paragraph(_money_markup(tot_out, cur), st['tf_r']),
            Paragraph(_money_markup(tot_in - tot_out, cur), st['tf_r']),
            Paragraph(_money_markup(closing, cur), st['tf_r']),
        ])
        col_w = [_USABLE_W - 4 * 100, 100, 100, 100, 100]
        t = Table(table, colWidths=col_w, repeatRows=1)
        n = len(table)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), Brand.BAND),
            ('LINEBELOW', (0, 0), (-1, 0), 1.5, Brand.INK),
            ('ROWBACKGROUNDS', (0, 1), (-1, n - 2), [colors.white, Brand.ZEBRA]),
            ('LINEABOVE', (0, n - 1), (-1, n - 1), 1.5, Brand.INK),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))
        if len(data) > 1:
            story.append(Paragraph(f"Currency: {cur}", st['biz_name']))
            story.append(Spacer(1, 4))
        story.append(_pl_line('Opening cash balance', opening, cur, st, 'item'))
        story.append(Spacer(1, 6))
        story.append(t)
        story.append(Spacer(1, 6))
        story.append(_pl_line('Closing cash balance', closing, cur, st, 'grand'))
        story.append(Spacer(1, 12))
    if not data:
        story.append(Paragraph('No cash movements found for this period.', st['note']))
    return story


def _render_pdf(kind, data, meta):
    tmpdir = _tmpdir(meta)
    title = meta['title']
    path = os.path.join(tmpdir, _slug(title) + '.pdf')
    st = _styles()

    doc = SimpleDocTemplate(
        path, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm, topMargin=18 * mm, bottomMargin=22 * mm,
        title=title, author='TaLi',
    )

    story = _letterhead(title, meta, st) + _bizstrip(meta, st)
    if kind == 'income_statement':
        story += _income_story(data, meta, st)
    elif kind == 'cashflow':
        story += _cashflow_story(data, meta, st)
    else:
        story += _account_story(data, meta, st)

    business = meta.get('business_name') or 'TaLi'
    footer = _make_footer(title, business)
    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    return {'path': path, 'filename': os.path.basename(path), 'mime': PDF_MIME}


def _letterhead(title, meta, st):
    """Wordmark + tagline (left) · doc title + period + generated (right) + terracotta rule."""
    wordmark = Paragraph(
        f'<font name="{F_DISPLAY}" size="30">TaLi<font color="#C2562F">.</font></font>',
        ParagraphStyle('wm', fontName=F_DISPLAY, fontSize=30, leading=32, textColor=Brand.INK))
    tagline = Paragraph('BOOKKEEPING IN YOUR CHAT',
                        ParagraphStyle('tag', fontName=F_BODY, fontSize=8.5, leading=12,
                                       textColor=Brand.SOFT, spaceBefore=3))
    period = meta.get('subtitle') or 'All time'
    gen = datetime.now().strftime('%d %b %Y')
    right = Paragraph(
        f'<font name="{F_DISPLAY}" size="16" color="#241F1A">{title}</font><br/>'
        f'<font size="10">{period}</font><br/><font size="10">Generated {gen}</font>',
        ParagraphStyle('dm', fontName=F_BODY, fontSize=10, leading=15, textColor=Brand.SOFT, alignment=2))
    head = Table([[[wordmark, tagline], right]], colWidths=[_USABLE_W * 0.55, _USABLE_W * 0.45])
    head.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0), ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0), ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    return [head, HRFlowable(width='100%', thickness=2, color=Brand.ACCENT, spaceBefore=10, spaceAfter=2)]


def _bizstrip(meta, st):
    business = meta.get('business_name') or 'TaLi'
    addr = meta.get('business_address') or ''
    left = Paragraph(f'<font name="{F_DISPLAY}" size="14" color="#241F1A">{business}</font>'
                     + (f'<br/><font size="10">{addr}</font>' if addr else ''),
                     ParagraphStyle('bl', fontName=F_BODY, fontSize=10, leading=16, textColor=Brand.SOFT))
    right = Paragraph(meta.get('strip_right') or 'Prepared by TaLi', st['soft_r'])
    strip = Table([[left, right]], colWidths=[_USABLE_W * 0.6, _USABLE_W * 0.4])
    strip.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0), ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 12), ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    return [strip]


def _make_footer(doc_title, business):
    def _f(canvas, doc):
        canvas.saveState()
        w, h = A4
        m = 18 * mm
        canvas.setStrokeColor(Brand.LINE_SOFT)
        canvas.setLineWidth(0.5)
        canvas.line(m, 14 * mm, w - m, 14 * mm)
        canvas.setFont(F_BODY, 7.5)
        canvas.setFillColor(Brand.SOFT)
        canvas.drawString(m, 10 * mm, f"TaLi · {doc_title} · {business}")
        canvas.setFillColor(Brand.ACCENT)
        canvas.drawCentredString(w / 2, 10 * mm, "Every entry validated & auditable")
        canvas.setFillColor(Brand.SOFT)
        canvas.drawRightString(w - m, 10 * mm, f"Page {doc.page}")
        canvas.restoreState()
    return _f


def _short_date(date_str):
    try:
        return datetime.strptime(str(date_str)[:10], '%Y-%m-%d').strftime('%d %b')
    except Exception:
        return str(date_str)[:10]


def _month_label(month_str):
    try:
        return datetime.strptime(str(month_str), '%Y-%m').strftime('%b %Y')
    except Exception:
        return str(month_str)


# ============================================================ XLSX

def _render_xlsx(kind, data, meta):
    from openpyxl import Workbook

    tmpdir = _tmpdir(meta)
    path = os.path.join(tmpdir, _slug(meta['title']) + '.xlsx')
    wb = Workbook()
    wb.remove(wb.active)

    if kind == 'income_statement':
        _xlsx_summary(wb, compute_income_statement(data), meta)
        _xlsx_transactions(wb, _group_rows_by_currency(data), meta)
    elif kind == 'cashflow':
        _xlsx_cashflow(wb, data, meta)
    else:
        _xlsx_transactions(wb, _group_rows_by_currency(data), meta)

    if not wb.sheetnames:
        ws = wb.create_sheet('Statement')
        ws['A1'] = 'No records found for this period.'
    wb.save(path)
    return {'path': path, 'filename': os.path.basename(path), 'mime': XLSX_MIME}


def _money_fmt(currency):
    sym = '₦' if (currency or 'NGN').upper() == 'NGN' else _CCY.get((currency or '').upper(), '')
    return f'"{sym}"#,##0.00;[Red]("{sym}"#,##0.00)'


def _xlsx_transactions(wb, by_cur, meta):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_fill = PatternFill('solid', fgColor='1E6E43')
    header_font = Font(color='FFFFFF', bold=True)
    bold = Font(bold=True)
    thin_top = Border(top=Side(style='thin', color='241F1A'))
    cols = ['Date', 'Description', 'Category', 'Type', 'Money in', 'Money out', 'Balance']
    for cur, rows in sorted(by_cur.items()):
        title = ('Transactions' if len(by_cur) == 1 else f'Transactions {cur}')[:31]
        ws = wb.create_sheet(title=title)
        ws.append(cols)
        mfmt = _money_fmt(cur)
        balance = 0.0
        for r in rows:
            amt = float(r['amount'])
            is_in = r.get('type') == 'income'
            balance += amt if is_in else -amt
            desc = (r.get('item') or (r.get('action') or '').title() or '')
            if r.get('action'):
                desc = f"{desc} — {r['action']}"
            row = [str(r['date'])[:10], desc, r.get('category') or '', (r.get('type') or '').title(),
                   amt if is_in else None, amt if not is_in else None, None]
            ws.append(row)
            rn = ws.max_row
            # live running balance: (prev balance +) money in − money out
            if rn == 2:
                ws[f'G{rn}'] = f'=IFERROR(E{rn},0)-IFERROR(F{rn},0)'
            else:
                ws[f'G{rn}'] = f'=G{rn-1}+IFERROR(E{rn},0)-IFERROR(F{rn},0)'
        last = ws.max_row
        # totals row with live formulas
        trow = last + 1
        ws.cell(row=trow, column=1, value='Totals')
        ws.cell(row=trow, column=5, value=f'=SUM(E2:E{last})')
        ws.cell(row=trow, column=6, value=f'=SUM(F2:F{last})')
        ws.cell(row=trow, column=7, value=(f'=G{last}' if last >= 2 else 0))
        # number formats + alignment
        for rn in range(2, trow + 1):
            for cn in (5, 6, 7):
                c = ws.cell(row=rn, column=cn)
                c.number_format = mfmt
                c.alignment = Alignment(horizontal='right')
        # header styling
        for cn in range(1, len(cols) + 1):
            c = ws.cell(row=1, column=cn)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal='center')
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:G{last}'
        for cn in range(1, len(cols) + 1):
            ws.cell(row=trow, column=cn).font = bold
            ws.cell(row=trow, column=cn).border = thin_top
        widths = [12, 40, 16, 10, 15, 15, 16]
        for i, wdt in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = wdt


def _xlsx_summary(wb, by_cur, meta):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    bold = Font(bold=True)
    head = Font(bold=True, size=12)
    accent_fill = PatternFill('solid', fgColor='EAF3EC')
    thin_top = Border(top=Side(style='thin', color='241F1A'))
    ws = wb.create_sheet(title='Summary')
    ws['A1'] = (meta.get('business_name') or 'TaLi')
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Income Statement — {meta.get('subtitle') or 'all time'}"
    ws['A2'].font = Font(italic=True, color='6B6258')

    r = 4
    for cur, p in sorted(by_cur.items()):
        mfmt = _money_fmt(cur)

        def line(label, amount=None, *, bold_=False, top=False, fill=False, indent=0):
            nonlocal r
            a = ws.cell(row=r, column=1, value=label)
            a.alignment = Alignment(indent=indent)
            if amount is not None:
                b = ws.cell(row=r, column=2, value=amount)
                b.number_format = mfmt
                b.alignment = Alignment(horizontal='right')
                if bold_:
                    b.font = bold
                if top:
                    b.border = thin_top
                if fill:
                    b.fill = accent_fill
            if bold_:
                a.font = bold
            if top:
                a.border = thin_top
            if fill:
                a.fill = accent_fill
            r += 1

        if len(by_cur) > 1:
            line(f"Currency: {cur}", bold_=True)
        line('Revenue', bold_=True)
        for label, amt in (p['revenue'] or [('Sales', 0.0)]):
            line(label, amt, indent=1)
        line('Total revenue', p['total_revenue'], bold_=True, top=True)
        r += 1
        line('Cost of goods sold', bold_=True)
        line('Purchases', p['purchases'], indent=1)
        line('Cost of goods sold', p['cogs'], bold_=True, top=True)
        line('Gross profit', p['gross_profit'], bold_=True, top=True)
        r += 1
        line('Operating expenses', bold_=True)
        for label, amt in (p['expenses'] or []):
            line(label, amt, indent=1)
        line('Total operating expenses', p['total_expenses'], bold_=True, top=True)
        r += 1
        line('Net profit', p['net_profit'], bold_=True, top=True, fill=True)
        r += 2
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 18


def _xlsx_cashflow(wb, data, meta):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_fill = PatternFill('solid', fgColor='1E6E43')
    header_font = Font(color='FFFFFF', bold=True)
    bold = Font(bold=True)
    thin_top = Border(top=Side(style='thin', color='241F1A'))
    cols = ['Month', 'Cash in', 'Cash out', 'Net', 'Balance']
    for cur, rows in sorted((data or {}).items()):
        title = ('Cashflow' if len(data) == 1 else f'Cashflow {cur}')[:31]
        ws = wb.create_sheet(title=title)
        ws.append(cols)
        mfmt = _money_fmt(cur)
        for rec in rows:
            ws.append([_month_label(rec['month']), rec['inflow'], rec['outflow'], rec['net'], rec['cumulative']])
        last = ws.max_row
        trow = last + 1
        ws.cell(row=trow, column=1, value='Totals')
        ws.cell(row=trow, column=2, value=f'=SUM(B2:B{last})')
        ws.cell(row=trow, column=3, value=f'=SUM(C2:C{last})')
        ws.cell(row=trow, column=4, value=f'=SUM(D2:D{last})')
        ws.cell(row=trow, column=5, value=(f'=E{last}' if last >= 2 else 0))
        for rn in range(2, trow + 1):
            for cn in (2, 3, 4, 5):
                c = ws.cell(row=rn, column=cn)
                c.number_format = mfmt
                c.alignment = Alignment(horizontal='right')
        for cn in range(1, len(cols) + 1):
            c = ws.cell(row=1, column=cn)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal='center')
            ws.cell(row=trow, column=cn).font = bold
            ws.cell(row=trow, column=cn).border = thin_top
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:E{last}'
        for i, wdt in enumerate([14, 16, 16, 16, 16], start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = wdt
